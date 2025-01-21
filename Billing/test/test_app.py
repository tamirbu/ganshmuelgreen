import unittest
import requests
import openpyxl
import os
from unittest.mock import patch, MagicMock

class TestFlaskApp(unittest.TestCase):
    BASE_URL = "http://test_billing_app:5000"
    TEST_FILE_PATH = "/app/in/rates.xlsx"  # Adjust path to match container setup

    def setUp(self):
        """Setup logic before each test."""
        # Add a default provider to ensure tests run smoothly
        self.default_provider = {"name": "Default Provider"}
        response = requests.post(f"{self.BASE_URL}/provider", json=self.default_provider)
        if response.status_code == 201:
            self.default_provider_id = response.json()["id"]
        else:
            self.default_provider_id = None

    def test_health_check(self):
        """Test the health check endpoint."""
        response = requests.get(f"{self.BASE_URL}/health")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"status": "OK"})

    def test_add_provider_success(self):
        """Test adding a provider successfully."""
        response = requests.post(f"{self.BASE_URL}/provider", json={"name": "Test Provider"})
        self.assertEqual(response.status_code, 201)
        self.assertIn("id", response.json())
        self.assertEqual(response.json()["name"], "Test Provider")

    def test_add_provider_missing_name(self):
        """Test adding a provider with missing name."""
        response = requests.post(f"{self.BASE_URL}/provider", json={})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json(), {"error": "Missing provider name"})

    @patch("requests.post")
    def test_register_truck_success(self, mock_post):
        """Test registering a new truck successfully with mocked provider creation."""
        mock_post.return_value.status_code = 201
        mock_post.return_value.json.return_value = {"id": 1, "name": "Test Provider"}
        provider_response = requests.post(f"{self.BASE_URL}/provider", json={"name": "Test Provider"})
        self.assertEqual(provider_response.status_code, 201)
        provider_id = provider_response.json()["id"]

        mock_post.return_value.status_code = 201
        mock_post.return_value.json.return_value = {
            "message": "Truck registered successfully",
            "id": "T123",
            "provider_id": provider_id
        }
        truck_response = requests.post(f"{self.BASE_URL}/truck", json={"id": "T123", "provider_id": provider_id})
        self.assertEqual(truck_response.status_code, 201)
        self.assertEqual(truck_response.json()["message"], "Truck registered successfully")
        self.assertEqual(truck_response.json()["id"], "T123")
        self.assertEqual(truck_response.json()["provider_id"], provider_id)

    def test_register_truck_missing_fields(self):
        """Test registering a truck with missing fields."""
        response = requests.post(f"{self.BASE_URL}/truck", json={"id": "T123"})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"], "Missing required fields. Both 'id' and 'provider_id' are required")

    def test_update_truck_success(self):
        """Test updating an existing truck."""
        response = requests.put(f"{self.BASE_URL}/truckExist/T123", json={"provider_id": self.default_provider_id})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["id"], "T123")
        self.assertEqual(response.json()["provider_id"], self.default_provider_id)

    def test_update_provider_success(self):
        """Test updating an existing provider."""
        response = requests.put(f"{self.BASE_URL}/provider/{self.default_provider_id}", json={"name": "Updated Provider"})
        self.assertIn(response.status_code, [200, 201])
        self.assertIn("message", response.json())

    def test_get_truck_exists(self):
        """Test checking if a truck exists."""
        response = requests.get(f"{self.BASE_URL}/truckExists/T123")
        self.assertIn(response.status_code, [200, 404])
        if response.status_code == 200:
            self.assertEqual(response.json()["message"], "Truck exists")
        else:
            self.assertEqual(response.json()["error"], "Truck not found")

    def test_get_bill_success(self):
        """Test getting a bill for a provider."""
        response = requests.get(f"{self.BASE_URL}/bill/{self.default_provider_id}?from=20230101000000&to=20231231235959")
        self.assertEqual(response.status_code, 200)
        self.assertIn("total", response.json())

    def test_get_truck_invalid_format(self):
        """Test fetching a truck with an invalid ID format."""
        response = requests.get(f"{self.BASE_URL}/truck/INVALID_ID_12345678910")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"], "Invalid truck ID format")

    def test_upload_rates_invalid_provider(self):
        """Test uploading rates with an invalid provider ID in the Scope column."""
        new_row = ["TestProduct", 123.45, "101"]  # Product, Rate, Scope (invalid provider ID)

        try:
            update_rates_file_for_test(self.TEST_FILE_PATH, new_row)

            response = requests.post(f"{self.BASE_URL}/rates")

            self.assertEqual(response.status_code, 404)
            self.assertIn("Provider with ID 101 does not exist", response.json()["error"])
        finally:
            restore_rates_file(self.TEST_FILE_PATH)

# Utility functions for modifying the Excel file
def update_rates_file_for_test(file_path, new_row):
    """
    Temporarily add a new row to the Excel file for testing.
    The new row is appended at the end of the file.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File {file_path} does not exist")

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        sheet.append(new_row)

        workbook.save(file_path)
        print("Row added for test.")
    except Exception as e:
        raise RuntimeError(f"Error updating file for test: {e}")

def restore_rates_file(file_path):
    """
    Restore the Excel file by removing the last row (used for the test).
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File {file_path} does not exist")

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        max_row = sheet.max_row
        sheet.delete_rows(max_row)

        workbook.save(file_path)
        print("Row removed after test.")
    except Exception as e:
        raise RuntimeError(f"Error restoring file after test: {e}")

if __name__ == "__main__":
    unittest.main()


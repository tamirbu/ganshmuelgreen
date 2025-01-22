from flask import Flask, request, jsonify, send_file
import mysql.connector
import os
from openpyxl import load_workbook
import requests
from datetime import datetime



app = Flask(__name__)

# Base URL of the external service
container_name = os.getenv('CONTAINER_NAME')
WEIGHT_APP_URL = f"http://{container_name}:5000"

#db connaction-------------------------------------------------------------------------------------
db_host = os.getenv("DATABASE_HOST", "localhost")
db_port = os.getenv("DATABASE_PORT", "3306")
db_name = os.getenv("DATABASE_NAME", "mydatabase")
db_user = os.getenv("DATABASE_USER", "myuser")
db_password = os.getenv("DATABASE_PASSWORD", "mypassword")

def get_db_connection():
    return mysql.connector.connect(
        host=db_host,
        port=db_port,
        database=db_name,
        user=db_user,
        password=db_password
    )

# post------------------------------------------------------------------------------------------------------
@app.route("/provider", methods=["POST"])
def add_provider():
    # get the name
    data = request.get_json()

    if not data or "name" not in data:
        return jsonify({f"error": "Missing provider name"}), 400

    provider_name = data["name"]

    try:
        # open the connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # add provider to table
        query = "INSERT INTO Provider (name) VALUES (%s)"
        cursor.execute(query, (provider_name,))
        conn.commit()

        # get the id random
        provider_id = cursor.lastrowid

        # close connaction
        cursor.close()
        conn.close()

        # return json new provider
        return jsonify({"id": provider_id, "name": provider_name}), 201

    except mysql.connector.Error as err:
        return jsonify({"error": str(err)}), 500


@app.route("/rates", methods=["POST"])
def upload_rates():
    file_path = "/app/in/rates.xlsx"

    # Check if the file exists and is accessible
    if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
        print(f"File {file_path} not found or not accessible")
        return jsonify({"error": f"File {file_path} not found or not accessible"}), 404

    try:
        # Load workbook using openpyxl
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=file_path)
            ws = wb.active
        except ValueError as ve:
            print(f"Value error: {ve}")
            return jsonify({"error": "Invalid file format"}), 400
        except Exception as e:
            print(f"Unexpected error while loading file: {e}")
            return jsonify({"error": "Failed to process file"}), 500

        # Open MySQL connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # Process rows
        rows = list(ws.rows)[1:]  # Skip header row
        for row in rows:
            try:
                product = str(row[0].value)
                rate = float(row[1].value)
                scope = str(row[2].value)

                if scope.upper() != "ALL":
                    cursor.execute("SELECT id FROM Provider WHERE id = %s", (scope,))
                    provider_exists = cursor.fetchone()
                    if not provider_exists:
                        print(f"Provider {scope} not found")
                        return jsonify({"error": f"Provider with ID {scope} does not exist"}), 404

                if scope.upper() == "ALL":
                    cursor.execute(
                        "DELETE FROM Rates WHERE product_id = %s AND (scope IS NULL OR scope = 'ALL')",
                        (product,)
                    )
                    cursor.execute(
                        "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, 'ALL')",
                        (product, rate)
                    )
                else:
                    cursor.execute(
                        "DELETE FROM Rates WHERE product_id = %s AND scope = %s",
                        (product, scope)
                    )
                    cursor.execute(
                        "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, %s)",
                        (product, rate, scope)
                    )

                conn.commit()
            except ValueError as ve:
                print(f"Value error in row: {ve}")
                return jsonify({"error": f"Invalid data format: {ve}"}), 400
            except mysql.connector.Error as me:
                print(f"Database error: {me}")
                return jsonify({"error": f"Database error: {me}"}), 500

        # Close resources
        cursor.close()
        conn.close()
        wb.close()

        return jsonify({"message": "Rates updated successfully!"}), 200

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        return jsonify({"error": str(e)}), 500

    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/truck", methods=["POST"])
def register_truck():
    """
    Registers a new truck in the system.
    
    Expected JSON body:
    {
        "id": "license-plate",
        "provider_id": provider-id
    }
    
    Returns:
        201: Successfully registered truck
        400: Missing or invalid parameters
        404: Provider not found
        500: Database error
    """
    # Get JSON data from request
    data = request.get_json()
    
    # Validate required fields
    if not data or "id" not in data or "provider_id" not in data:
        return jsonify({
            "error": "Missing required fields. Both 'id' and 'provider_id' are required"
        }), 400
        
    truck_id = data["id"]
    provider_id = data["provider_id"]
    
    try:
        # Establish database connection
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if provider exists
        cursor.execute("SELECT id FROM Provider WHERE id = %s", (provider_id,))
        provider_exists = cursor.fetchone()
        
        if not provider_exists:
            cursor.close()
            conn.close()
            return jsonify({
                "error": f"Provider with ID {provider_id} does not exist"
            }), 404
            
        # Check if truck already exists
        cursor.execute("SELECT id FROM Trucks WHERE id = %s", (truck_id,))
        existing_truck = cursor.fetchone()
        
        if existing_truck:
            cursor.close()
            conn.close()
            return jsonify({
                "error": f"Truck with ID {truck_id} already exists"
            }), 400
            
        # Insert new truck
        cursor.execute(
            "INSERT INTO Trucks (id, provider_id) VALUES (%s, %s)",
            (truck_id, provider_id)
        )
        
        # Commit the transaction
        conn.commit()
        
        # Close database connection
        cursor.close()
        conn.close()
        
        # Return success response
        return jsonify({
            "message": "Truck registered successfully",
            "id": truck_id,
            "provider_id": provider_id
        }), 201
        
    except mysql.connector.Error as err:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
        return jsonify({"error": str(err)}), 500

# put-------------------------------------------------------------------------------------------------------
@app.route("/truck/<id>", methods=["PUT"])
def update_truck(id):
    """
    Updates or creates a truck entry in the database.

    Parameters:
        id (str): The truck ID provided in the URL.
        JSON body:
            - provider_id (int): The provider ID to associate with the truck.

    Returns:
        JSON response:
            - On success (200): {"id": truck_id, "provider_id": provider_id}
            - On missing provider_id in request body (400): {"error": "Missing provider_id"}
            - On non-existent provider_id (404): {"error": "Provider ID does not exist"}
            - On server/database error (500): {"error": <error_message>}
    """
    # Parse the JSON body of the request
    data = request.get_json()

    # Validate that provider_id is present in the request body
    if not data or "provider_id" not in data:
        return jsonify({"error": "Missing provider_id"}), 400

    provider_id = data["provider_id"]

    try:
        # Establish a database connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if the provider exists in the Provider table
        cursor.execute("SELECT COUNT(*) FROM Provider WHERE id = %s", (provider_id,))
        provider_exists = cursor.fetchone()[0]

        # If the provider does not exist, return a 404 error
        if not provider_exists:
            return jsonify({"error": "Provider ID does not exist"}), 404

        # Check if the truck exists in the Trucks table
        cursor.execute("SELECT COUNT(*) FROM Trucks WHERE id = %s", (id,))
        truck_exists = cursor.fetchone()[0]

        if truck_exists:
            # Update the existing truck's provider_id
            query = "UPDATE Trucks SET provider_id = %s WHERE id = %s"
            cursor.execute(query, (provider_id, id))
        else:
            # Insert a new truck record
            query = "INSERT INTO Trucks (id, provider_id) VALUES (%s, %s)"
            cursor.execute(query, (id, provider_id))

        # Commit the transaction to the database
        conn.commit()

        # Close the database connection
        cursor.close()
        conn.close()

        # Return a success response
        return jsonify({"id": id, "provider_id": provider_id}), 200

    except mysql.connector.Error as err:
        # Handle database errors and return an error response
        return jsonify({"error": str(err)}), 500

#aviv
@app.route("/provider/<int:id>", methods=["PUT"])
def update_provider(id):
    conn = None
    try:
        # Get JSON data for provider update
        data = request.get_json()
        new_name = data.get("name")
        if new_name is None:
            return jsonify({"error": "Name is required"}), 400

        # Establish database connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check existing provider by ID
        cursor.execute("SELECT name FROM Provider WHERE id = %s", (id,))
        existing_provider = cursor.fetchone()

        if existing_provider:
            # If the provider exists, update the name
            cursor.execute("UPDATE Provider SET name = %s WHERE id = %s", (new_name, id))
            conn.commit()
            return jsonify({
                "message": "Provider updated",
                "id": id,
                "previous_name": existing_provider[0],
                "new_name": new_name
            }), 200  # Return 200 OK for an update

        else:
            # If the provider doesn't exist, insert a new provider
            cursor.execute("INSERT INTO Provider (id, name) VALUES (%s, %s)", (id, new_name))
            conn.commit()
            return jsonify({
                "message": "Provider created",
                "id": id,
                "name": new_name
            }), 201  # Return 201 Created for new provider

    except Exception as e:
        # Handle any errors
        if conn:
            conn.rollback()
        return jsonify({
            "error": "Operation failed",
            "details": str(e)
        }), 500

    finally:
        # Ensure resources are properly closed
        if 'cursor' in locals() and cursor:
            cursor.close()
        if conn:
            conn.close()

# get-------------------------------------------------------------------------------------------------------
@app.route("/truckExists/<id>", methods=["GET"])
def get_truck_exists(id):
    """
    Checks if a truck with the given ID exists in the database.

    Parameters:
        id (str): The truck ID provided in the URL.

    Returns:
        JSON response:
            - On success (200): {"message": "Truck exists", "id": truck_id}
            - On not found (404): {"error": "Truck not found"}
    """
    try:
        # Establish a database connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if the truck exists in the Trucks table
        query = "SELECT COUNT(*) FROM Trucks WHERE id = %s"
        cursor.execute(query, (id,))
        truck_exists = cursor.fetchone()[0]

        # If the truck exists, return success
        if truck_exists:
            cursor.close()
            conn.close()
            return jsonify({"message": "Truck exists", "id": id}), 200

        # If the truck does not exist, return not found
        else:
            cursor.close()
            conn.close()
            return jsonify({"error": "Truck not found"}), 404

    except mysql.connector.Error as err:
        # Handle database errors
        return jsonify({"error": str(err)}), 500
        
# #aviv
@app.route("/rates", methods=["GET"])
def get_rates():
    file_path = '/app/in/rates.xlsx'
    # Check if file exists
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404

    try:
        # Return the file for download
        return send_file(
            file_path,
            as_attachment=True,
            download_name="rates.xlsx",  # Suggested file name for download
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # Correct MIME type for Excel files
        )
    except Exception as e:
        print(f"Error: {str(e)}")  # Debug print
        return jsonify({"error": "Failed to process rates", "details": str(e)}), 500


def get_default_dates():
    """Returns the default t1 (start of the month) and t2 (current time) as strings."""
    now = datetime.now()
    t1 = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    t2 = now
    return t1.strftime('%Y%m%d%H%M%S'), t2.strftime('%Y%m%d%H%M%S')

@app.route("/truck/<id>", methods=["GET"])
def get_truck(id):
    """
    Handles GET /truck/<id>?from=t1&to=t2 endpoint.

    Parameters:
        id (str): The truck ID.

    Query Parameters:
        from (str): Start datetime (yyyymmddhhmmss).
        to (str): End datetime (yyyymmddhhmmss).

    Returns:
        JSON response with truck details or an error message.
    """
    # Extract query parameters
    t1 = request.args.get("from")
    t2 = request.args.get("to")

    # If t1 or t2 are missing, set default values
    if not t1 or not t2:
        default_t1, default_t2 = get_default_dates()
        t1 = t1 or default_t1
        t2 = t2 or default_t2
        print(f"Using default dates: from={t1}, to={t2}")

    # Prepare query parameters for the external request
    params = {}
    if t1:
        params["from"] = t1
    if t2:
        params["to"] = t2

    # Validate the truck ID format (example: alphanumeric with dashes)
    if not isinstance(id, str) or len(id) > 10:
        return jsonify({"error": "Invalid truck ID format"}), 400

    try:
        # Call the external weight_app service
        response = requests.get(f"{WEIGHT_APP_URL}/item/{id}", params=params)

        # Handle response from weight_app
        if response.status_code == 404:
            return jsonify({"error": "Truck ID not found"}), 404

        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch data from weight service", "details": response.text}), 500

        # Parse the response from weight_app
        data = response.json()

        # Transform the data for the truck endpoint
        transformed_data = {
            "id": data["id"],
            "tara": data.get("tara", "na"),
            "sessions": data.get("sessions", [])
        }

        return jsonify(transformed_data), 200

    except requests.exceptions.RequestException as e:
        # Handle network or request errors
        print(f"Failed to connect to weight service: {e}")
        return jsonify({"error": "Failed to connect to weight service", "details": str(e)}), 500


@app.route("/bill/<id>", methods=["GET"])
def get_bill(id):
    """
    Get billing information for a provider within a specified time range.
    """
    try:
        # Parse date parameters
        t1 = request.args.get("from")
        t2 = request.args.get("to")
        now = datetime.now()
        if not t2:
            t2 = now.strftime("%Y%m%d%H%M%S")
        if not t1:
            t1 = now.replace(day=1, hour=0, minute=0, second=0).strftime("%Y%m%d%H%M%S")
        # Get provider information
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name FROM Provider WHERE id = %s", (id,))
        provider = cursor.fetchone()
        if not provider:
            return jsonify({"error": f"Provider {id} not found"}), 404
        # Get all trucks for this provider
        cursor.execute("SELECT id FROM Trucks WHERE provider_id = %s", (id,))
        provider_trucks = {truck["id"] for truck in cursor.fetchall()}
        # Get rates for this provider
        cursor.execute(
            """
            SELECT product_id, rate, scope
            FROM Rates
            WHERE scope = %s OR scope = 'ALL'
            ORDER BY scope DESC
            """,
            (id,)
        )
        rates = cursor.fetchall()
        # Create rates lookup dictionary
        rates_lookup = {rate['product_id']: rate['rate'] for rate in rates}
        cursor.close()
        conn.close()
        # Get sessions from weight service
        try:
            weight_response = requests.get(
                f"{WEIGHT_APP_URL}/weight",
                params={"from": t1, "to": t2},
                timeout=5
            )
            weight_response.raise_for_status()
            transactions = weight_response.json()
        except requests.RequestException as e:
            return jsonify({"error": f"Failed to fetch data from weight service: {str(e)}"}), 503
        # Process transactions and calculate billing
        products = {}
        session_count = 0
        missing_rates = set()
        for transaction in transactions:
            transaction_id = transaction.get("id")
            direction = transaction.get("direction")
            produce = transaction.get("produce")
            # check the out direction and not null productions
            if direction != "out" or produce == "na":
                continue
            try:
                # Get session details from /session/<id>
                session_response = requests.get(f"{WEIGHT_APP_URL}/session/{transaction_id}", timeout=5)
                session_response.raise_for_status()
                session = session_response.json()
            except requests.RequestException as e:
                print(f"Failed to fetch session {transaction_id}: {str(e)}")
                continue
            # get just the provider truck
            truck_id = session.get("truck")
            if truck_id not in provider_trucks:
                continue
            # session counter
            session_count += 1
            # check the produce rate
            if produce not in rates_lookup:
                missing_rates.add(produce)
                continue
            if produce not in products:
                products[produce] = {
                    "product": produce,
                    "count": 0,
                    "amount": 0,
                    "rate": rates_lookup[produce],
                    "pay": 0
                }
            products[produce]["count"] += 1
            if session.get('neto') and session['neto'] != 'na':
                amount = session['neto']
                products[produce]["amount"] += amount
                products[produce]["pay"] += (amount * rates_lookup[produce])
        # print the products that have no rate
        if missing_rates:
            return jsonify({
                "error": "Missing rates for the following products.",
                "missing_products": list(missing_rates),
                "message": "Please add rates for these products in the Rates table."
            }), 400
        # Calculate total payment
        total = sum(p["pay"] for p in products.values())
        # bill
        response = {
            "id": str(provider["id"]),
            "name": provider["name"],
            "from": format_date(t1),
            "to": format_date(t2),
            "truckCount": len(provider_trucks),
            "sessionCount": session_count,
            "products": list(products.values()),
            "total": total
        }
        return jsonify(response), 200
    except Exception as e:
        print(f"Error generating bill: {e}")
        return jsonify({"error": "Internal server error", "details": str(e)}), 500

def format_date(timestamp):
    """Converts YYYYMMDDHHMMSS to d/m/y format."""
    return datetime.strptime(timestamp, "%Y%m%d%H%M%S").strftime("%d/%m/%Y")



# home page--------------------------------------------------------------------------------------
@app.route("/")
def index():
    return "Welcome to the Flask App!"

# chack health------------------------------------------------------------------------------------
@app.route("/health")
def health_check():
    return jsonify({"status": "OK"}), 200

# main--------------------------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
